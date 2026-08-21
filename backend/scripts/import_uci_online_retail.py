"""Import UCI Online Retail II into a separate, traceable analysis table.

The published workbook represents anonymized real transactions.  It must never
be mapped into the teaching manufacturing tables because its customer, country,
and stock-code fields do not mean supplier, factory, or production order.
"""

from __future__ import annotations

import os
import tempfile
import time
import urllib.request
from datetime import datetime
from decimal import Decimal, InvalidOperation
from http.client import IncompleteRead
from pathlib import Path

import psycopg
from openpyxl import load_workbook

URL = "https://archive.ics.uci.edu/ml/machine-learning-databases/00502/online_retail_II.xlsx"
DATABASE_URL = os.getenv(
    "ONLINE_RETAIL_DATABASE_URL",
    "postgresql://supplymind_ro:supplymind-demo-ro@retail-postgres:5432/online_retail",
)
EXPECTED_COLUMNS = (
    "Invoice",
    "StockCode",
    "Description",
    "Quantity",
    "InvoiceDate",
    "Price",
    "Customer ID",
    "Country",
)
MINIMUM_ROWS = 1_000_000
BATCH_SIZE = 5_000


def download_workbook(destination: Path) -> None:
    request = urllib.request.Request(URL, headers={"User-Agent": "SupplyMind dataset importer"})
    failure: Exception | None = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(request, timeout=120) as response, destination.open("wb") as output:  # noqa: S310
                while chunk := response.read(1024 * 1024):
                    output.write(chunk)
            return
        except (IncompleteRead, OSError) as exc:
            failure = exc
            destination.unlink(missing_ok=True)
            if attempt < 2:
                time.sleep(2**attempt)
    raise RuntimeError("Unable to download the complete UCI Online Retail II workbook") from failure


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def as_decimal(value: object) -> Decimal:
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError) as exc:
        raise RuntimeError(f"Invalid numeric value in UCI Online Retail II: {value!r}") from exc


def rows_from_workbook(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    source_row_id = 0
    try:
        for worksheet in workbook.worksheets:
            headings = tuple(cell.value for cell in next(worksheet.iter_rows(max_row=1)))
            if headings != EXPECTED_COLUMNS:
                raise RuntimeError(f"Unexpected columns in worksheet {worksheet.title}: {headings!r}")
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not any(value is not None for value in row):
                    continue
                source_row_id += 1
                invoice_no, stock_code, description, quantity, invoice_at, unit_price, customer_id, country = row
                if not isinstance(invoice_at, datetime):
                    raise RuntimeError(f"Invalid invoice date at source row {source_row_id}: {invoice_at!r}")
                if not (invoice_no and stock_code and country):
                    raise RuntimeError(f"Missing required transaction field at source row {source_row_id}")
                yield (
                    source_row_id,
                    normalize_text(invoice_no),
                    normalize_text(stock_code),
                    normalize_text(description),
                    as_decimal(quantity),
                    invoice_at,
                    as_decimal(unit_price),
                    normalize_text(customer_id),
                    normalize_text(country),
                )
    finally:
        workbook.close()


def ensure_table(cursor: psycopg.Cursor) -> None:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS retail_transactions (
          source_row_id INTEGER PRIMARY KEY,
          invoice_no TEXT NOT NULL,
          stock_code TEXT NOT NULL,
          description TEXT,
          quantity NUMERIC(16, 2) NOT NULL,
          invoice_at TIMESTAMP NOT NULL,
          unit_price NUMERIC(16, 4) NOT NULL,
          customer_id TEXT,
          country TEXT NOT NULL,
          dataset_source TEXT NOT NULL DEFAULT 'UCI Online Retail II',
          imported_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS retail_transactions_invoice_at_idx
          ON retail_transactions (invoice_at);
        CREATE INDEX IF NOT EXISTS retail_transactions_stock_code_idx
          ON retail_transactions (stock_code);
        CREATE INDEX IF NOT EXISTS retail_transactions_country_idx
          ON retail_transactions (country);
        """
    )


def import_rows(path: Path) -> int:
    inserted = 0
    batch: list[tuple] = []
    with psycopg.connect(DATABASE_URL) as connection:
        with connection.cursor() as cursor:
            ensure_table(cursor)
            for row in rows_from_workbook(path):
                batch.append(row)
                if len(batch) >= BATCH_SIZE:
                    cursor.executemany(
                        """
                        INSERT INTO retail_transactions
                          (source_row_id, invoice_no, stock_code, description, quantity,
                           invoice_at, unit_price, customer_id, country)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (source_row_id) DO UPDATE SET
                          invoice_no = EXCLUDED.invoice_no,
                          stock_code = EXCLUDED.stock_code,
                          description = EXCLUDED.description,
                          quantity = EXCLUDED.quantity,
                          invoice_at = EXCLUDED.invoice_at,
                          unit_price = EXCLUDED.unit_price,
                          customer_id = EXCLUDED.customer_id,
                          country = EXCLUDED.country
                        """,
                        batch,
                    )
                    inserted += len(batch)
                    batch.clear()
            if batch:
                cursor.executemany(
                    """
                    INSERT INTO retail_transactions
                      (source_row_id, invoice_no, stock_code, description, quantity,
                       invoice_at, unit_price, customer_id, country)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (source_row_id) DO UPDATE SET
                      invoice_no = EXCLUDED.invoice_no,
                      stock_code = EXCLUDED.stock_code,
                      description = EXCLUDED.description,
                      quantity = EXCLUDED.quantity,
                      invoice_at = EXCLUDED.invoice_at,
                      unit_price = EXCLUDED.unit_price,
                      customer_id = EXCLUDED.customer_id,
                      country = EXCLUDED.country
                    """,
                    batch,
                )
                inserted += len(batch)
        if inserted < MINIMUM_ROWS:
            raise RuntimeError(f"UCI Online Retail II is unexpectedly small: {inserted} rows")
        connection.commit()
    return inserted


if __name__ == "__main__":
    with tempfile.TemporaryDirectory(prefix="supplymind-online-retail-") as directory:
        workbook_path = Path(directory) / "online_retail_II.xlsx"
        download_workbook(workbook_path)
        print(f"Imported {import_rows(workbook_path)} UCI Online Retail II transaction rows")
